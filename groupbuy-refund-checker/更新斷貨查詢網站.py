# -*- coding: utf-8 -*-
"""
更新斷貨查詢網站.py — 團購斷貨退款查詢網站的資料更新＋部署

做三件事：
  1. 讀當月「出貨明細」＋「斷貨單」，整理成每位會員的查詢資料
  2. 上傳到 Google 試算表（網站資料庫）的 查詢資料 分頁（用 gws CLI）
  3. 把 網站\ 資料夾部署到 Cloudflare Pages（用 wrangler）

用法：
  python 更新斷貨查詢網站.py                    # 自動抓最新月團，全部做
  python 更新斷貨查詢網站.py --month 2026-07_7月團
  python 更新斷貨查詢網站.py --no-deploy        # 只更新試算表，不部署網站
  python 更新斷貨查詢網站.py --no-sheet         # 只部署網站，不動試算表
  python 更新斷貨查詢網站.py --dry-run          # 只整理資料印統計，什麼都不上傳
"""
import argparse
import glob
import json
import os
import re
import subprocess
import sys
import unicodedata

sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(BASE)


def load_cfg():
    p = os.path.join(BASE, "設定.json")
    try:
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"[錯誤] 讀不了 設定.json（檔案不見或格式壞了？）: {e}", file=sys.stderr)
        sys.exit(1)


def load_xlsx(path):
    import openpyxl
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"[錯誤] 讀不了 {os.path.basename(path)}（檔案損毀或被 Excel 開著？請關閉 Excel 再重跑）: {e}", file=sys.stderr)
        sys.exit(1)


def find_month(month_arg):
    """找月團資料夾：--month 指定名稱，否則取 每月資料\ 下名稱最新、且有出貨明細的那個。"""
    monthly = os.path.join(ROOT, "每月資料")
    if month_arg:
        p = month_arg if os.path.isdir(month_arg) else os.path.join(monthly, month_arg)
        if not os.path.isdir(p):
            print(f"[錯誤] 找不到月團資料夾：{month_arg}（找過 {p}；請確認 每月資料\\ 下的資料夾名稱）", file=sys.stderr)
            sys.exit(1)
        return p
    cands = [d for d in sorted(glob.glob(os.path.join(monthly, "????-??_*月團")), reverse=True)
             if find_detail_file(d, quiet=True)]
    if not cands:
        print(f"[錯誤] {monthly} 下找不到任何有「出貨明細」的月團資料夾；請先跑步驟5產生出貨明細，或用 --month 指定。", file=sys.stderr)
        sys.exit(1)
    return cands[0]


def find_detail_file(month_dir, quiet=False):
    """在 出貨單\ 夾找出貨明細主檔（排除 核對清單 與 ~$ 暫存）。"""
    files = [f for f in glob.glob(os.path.join(month_dir, "出貨單", "*出貨明細*.xlsx"))
             if "核對" not in os.path.basename(f) and not os.path.basename(f).startswith("~$")]
    if not files:
        if not quiet:
            print(f"[錯誤] {month_dir}\\出貨單\\ 裡找不到 *出貨明細*.xlsx，請先跑步驟5。", file=sys.stderr)
            sys.exit(1)
        return None
    return files[0]


def find_shortage_file(month_dir):
    files = [f for f in glob.glob(os.path.join(month_dir, "斷貨單", "*斷貨單*.xlsx"))
             if not os.path.basename(f).startswith("~$")]
    return files[0] if files else None


def norm(s):
    return unicodedata.normalize("NFKC", str(s)).strip() if s is not None else ""


def product_code(name):
    """從商品名稱抽出款式代碼：'1.mt01 韓女綁帶撞色上衣_白' -> 'mt01'。抽不到回傳空字串。"""
    t = norm(name)
    t = re.sub(r"^\d+\.\s*", "", t)
    m = re.match(r"([A-Za-z]+\d+)", t)
    return m.group(1).lower() if m else ""


def parse_refund_expr(v):
    """E欄退款金額：626 -> (626,[626])；'198+338=536' -> (536,[198,338])。看不懂回傳 (None,[])。"""
    if v is None:
        return None, []
    if isinstance(v, (int, float)):
        return round(v), [round(v)]
    t = norm(v).replace(" ", "")
    m = re.match(r"^([\d+.]+)=(\d+(?:\.\d+)?)$", t)
    if m:
        parts = [round(float(x)) for x in m.group(1).split("+") if x]
        return round(float(m.group(2))), parts
    if re.match(r"^\d+(?:\.\d+)?$", t):
        return round(float(t)), [round(float(t))]
    return None, []


def read_shipping_detail(path):
    """讀出貨明細：回傳 [ {id, name, total, shipTotal, feeProcess, feePickup, cod, items:[{n,p,q,f}]} ]（照檔案順序）。"""
    wb = load_xlsx(path)
    ws = wb.active
    members, cur = [], None
    for row in ws.iter_rows(min_row=2, values_only=True):
        a = norm(row[0]) if len(row) > 0 else ""
        g = norm(row[6]) if len(row) > 6 else ""
        c = row[2] if len(row) > 2 else None
        if a in ("會員",):
            continue  # 表頭
        if g and re.match(r"^M\d+$", g):
            cur = {
                "id": g,
                "name": a.split("_", 1)[1].strip() if "_" in a else a,
                "total": round(row[7] or 0),
                "shipTotal": round(row[9] or 0),
                "feeProcess": round(row[10] or 0),
                "feePickup": round(row[11] or 0),
                "cod": round(row[12] or 0),
                "items": [],
            }
            members.append(cur)
        elif not c:
            cur = None  # 分段標題列（例：7月團出貨單之一）或空白列
            continue
        if cur is not None and c:
            cur["items"].append({
                "n": re.sub(r"^\d+\.\s*", "", norm(c)),
                "p": round(row[3] or 0),
                "q": round(row[4] or 0),
                "f": round(row[8] or 0),
            })
    wb.close()
    return members


def read_shortage(path):
    """讀斷貨單。回傳 (shortages, weight_map)
    shortages: {會員編號: {"total": 退款總額, "items": [{"n":款式,"s":規格,"r":單筆退款或None}]}}
    weight_map: {款式代碼: 重量}（來自第一分頁）
    """
    wb = load_xlsx(path)
    # 第一分頁：商品名稱(D) -> 重量(I)
    weight_map = {}
    ws1 = wb.worksheets[0]
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if len(row) > 8 and row[3] and row[8] not in (None, ""):
            code = product_code(row[3])
            if code and code not in weight_map:
                try:
                    weight_map[code] = float(row[8])
                except (TypeError, ValueError):
                    pass
    # 退款分頁：名稱含「退款」或「斷貨」
    ws2 = None
    for nm in wb.sheetnames:
        if "退款" in nm or ("斷貨" in nm and wb[nm] is not ws1):
            ws2 = wb[nm]
            break
    shortages = {}
    if ws2 is None:
        wb.close()
        return shortages, weight_map
    groups = []  # [(id, total, parts, items)]
    cur = None
    for row in ws2.iter_rows(min_row=2, values_only=True):
        a = norm(row[0]) if len(row) > 0 else ""
        c = row[2] if len(row) > 2 else None
        d = norm(row[3]) if len(row) > 3 else ""
        e = row[4] if len(row) > 4 else None
        m = re.match(r"^(M\d+)", a)
        if m:
            total, parts = parse_refund_expr(e)
            cur = {"id": m.group(1), "total": total, "parts": parts, "items": []}
            groups.append(cur)
        if cur is not None and c:
            cur["items"].append({"n": re.sub(r"^\d+\.\s*", "", norm(c)), "s": d})
    wb.close()
    for gp in groups:
        items = gp["items"]
        parts = gp["parts"]
        # 金額算式的份數跟品項數一致時，逐項對應；不一致就只給總額
        for i, it in enumerate(items):
            it["r"] = parts[i] if len(parts) == len(items) else None
        if gp["total"] is None:
            print(f"[警告] 斷貨單裡 {gp['id']} 的退款金額看不懂（E欄），這位會員的應退金額會是 0，請檢查斷貨單。")
            gp["total"] = 0
        if len(parts) not in (0, len(items)):
            print(f"[警告] {gp['id']} 的退款算式份數({len(parts)})跟斷貨品項數({len(items)})對不上，網頁只顯示總退款額。")
        prev = shortages.get(gp["id"])
        if prev:
            print(f"[警告] 斷貨單裡 {gp['id']} 出現兩次，金額合併計算。")
            prev["total"] += gp["total"]
            prev["items"] += items
        else:
            shortages[gp["id"]] = {"total": gp["total"], "items": items}
    return shortages, weight_map


def read_name_fallback(month_dir):
    """從 全數訂單明細表 補會員姓名（給只在斷貨單、不在出貨明細的會員用）。"""
    files = [f for f in glob.glob(os.path.join(month_dir, "出貨單", "*全數訂單明細*.xlsx"))
             if not os.path.basename(f).startswith("~$")]
    names = {}
    if not files:
        return names
    wb = load_xlsx(files[0])
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        a = norm(row[0]) if row and len(row) > 0 else ""
        m = re.match(r"^(M\d+)_(.+)$", a)
        if m and m.group(1) not in names:
            names[m.group(1)] = m.group(2).strip()
    wb.close()
    return names


def build_rows(members, shortages, weight_map, name_fallback):
    """合併成試算表列。回傳 (rows, stats)。rows: [會員編號,姓名,總金額,二補加總,品項數,斷貨品項數,應退金額,明細JSON]"""
    rows, seen = [], {}
    for mb in members:
        if mb["id"] in seen:
            print(f"[警告] 出貨明細裡 {mb['id']} 出現兩次，只保留第一筆，請人工確認。")
            continue
        seen[mb["id"]] = True
        sh = shortages.get(mb["id"])
        out_items = list(sh["items"]) if sh else []
        # 把斷貨品項對回訂購清單：款式代碼相同＋規格吻合（規格x=不分規格）
        for oi in out_items:
            code, spec = product_code(oi["n"]), norm(oi.get("s", ""))
            hit = None
            for it in mb["items"]:
                if it.get("out"):
                    continue
                if product_code(it["n"]) != code:
                    continue
                if spec in ("", "x") or spec in it["n"] or it["n"].endswith(spec):
                    hit = it
                    break
            if hit is None:  # 規格對不上就退而求其次：同款式代碼即可
                for it in mb["items"]:
                    if not it.get("out") and product_code(it["n"]) == code:
                        hit = it
                        break
            if hit is not None:
                hit["out"] = True
                hit["r"] = oi.get("r")
                oi["matched"] = True
        unmatched = [oi for oi in out_items if not oi.get("matched")]
        for oi in unmatched:
            print(f"[提醒] {mb['id']} 的斷貨品項「{oi['n']} {oi.get('s','')}」在出貨明細找不到對應列，會另列在網頁的斷貨清單裡。")
        for it in mb["items"]:
            it["w"] = weight_map.get(product_code(it["n"]))
            it.setdefault("out", False)
        refund = sh["total"] if sh else 0
        payload = {
            "total": mb["total"], "shipTotal": mb["shipTotal"],
            "feeProcess": mb["feeProcess"], "feePickup": mb["feePickup"],
            "cod": mb["cod"], "refund": refund,
            "items": mb["items"],
            "extra": [{"n": oi["n"], "s": oi.get("s", ""), "r": oi.get("r")} for oi in unmatched],
        }
        rows.append([mb["id"], mb["name"], mb["total"], mb["shipTotal"], len(mb["items"]),
                     sum(1 for it in mb["items"] if it["out"]) + len(unmatched), refund,
                     json.dumps(payload, ensure_ascii=False, separators=(",", ":"))])
    # 只在斷貨單、不在出貨明細的會員（整單斷貨等情況）也要能查
    for sid, sh in shortages.items():
        if sid in seen:
            continue
        nm = name_fallback.get(sid, "")
        if not nm:
            print(f"[警告] {sid} 只出現在斷貨單、且查不到姓名，網站上需要姓名才能查詢——請人工確認這位會員。")
        payload = {"total": 0, "shipTotal": 0, "feeProcess": 0, "feePickup": 0, "cod": 0,
                   "refund": sh["total"], "items": [],
                   "extra": [{"n": it["n"], "s": it.get("s", ""), "r": it.get("r")} for it in sh["items"]]}
        rows.append([sid, nm, 0, 0, 0, len(sh["items"]), sh["total"],
                     json.dumps(payload, ensure_ascii=False, separators=(",", ":"))])
    stats = {
        "members": len(rows),
        "refund_members": sum(1 for r in rows if r[6]),
        "refund_total": sum(r[6] for r in rows),
    }
    return rows, stats


def _which(name):
    """Windows 上 gws/npx 是 .cmd 包裝檔，subprocess 要給完整路徑才找得到。"""
    import shutil
    p = shutil.which(name)
    if not p:
        print(f"[錯誤] 找不到 {name} 指令，請確認已安裝並在 PATH 裡。", file=sys.stderr)
        sys.exit(1)
    return p


def _gws_cmd():
    """優先直接用 node 呼叫 gws 底層腳本：.cmd 包裝檔經 cmd.exe 有 8191 字元上限，
    大批資料的 --json 會爆掉；直接 node 呼叫上限是 32767。"""
    gws = _which("gws")
    runjs = os.path.join(os.path.dirname(gws), "node_modules", "@googleworkspace", "cli", "run.js")
    node = __import__("shutil").which("node")
    if node and os.path.exists(runjs):
        return [node, runjs]
    return [gws]


def run_gws(args_list, body=None, params=None, allow_fail=False):
    cmd = _gws_cmd() + args_list
    if params is not None:
        cmd += ["--params", json.dumps(params, ensure_ascii=False)]
    if body is not None:
        cmd += ["--json", json.dumps(body, ensure_ascii=False)]
    flags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
    r = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace",
                       shell=False, creationflags=flags)
    if r.returncode != 0:
        if allow_fail:
            return None
        print(f"[錯誤] gws 指令失敗：{' '.join(args_list)}\n{(r.stderr or r.stdout)[:800]}", file=sys.stderr)
        sys.exit(1)
    return r.stdout


def upload_to_sheet(cfg, rows, month_label):
    sid = cfg["spreadsheetId"]
    sheet = cfg.get("查詢資料分頁", "查詢資料")
    print(f"[info] 清空 {sheet} 舊資料…")
    run_gws(["sheets", "spreadsheets", "values", "clear"],
            params={"spreadsheetId": sid, "range": f"{sheet}!A2:H"}, body={})
    # 分批上傳：每批 JSON 長度控制在 ~20000 字內（gws 只吃指令列參數，Windows 有長度上限）
    print(f"[info] 上傳 {len(rows)} 位會員資料…")
    batch, size, start_row, done = [], 0, 2, 0
    def flush():
        nonlocal batch, size, start_row, done
        if not batch:
            return
        run_gws(["sheets", "spreadsheets", "values", "update"],
                params={"spreadsheetId": sid, "range": f"{sheet}!A{start_row}",
                        "valueInputOption": "RAW"},
                body={"values": batch})
        done += len(batch)
        start_row += len(batch)
        print(f"  已上傳 {done}/{len(rows)}")
        batch, size = [], 0
    for row in rows:
        s = len(json.dumps(row, ensure_ascii=False))
        if batch and size + s > 20000:
            flush()
        batch.append(row)
        size += s
    flush()
    # 更新設定分頁的月團名稱（退款期限不動，由使用者自己在試算表改）
    run_gws(["sheets", "spreadsheets", "values", "update"],
            params={"spreadsheetId": sid, "range": f"{cfg.get('設定分頁','設定')}!B2",
                    "valueInputOption": "RAW"},
            body={"values": [[month_label]]})


def verify_sheet(cfg, rows):
    """上傳後抽查：讀回列數與退款總額，跟本地算的比對。"""
    sid = cfg["spreadsheetId"]
    sheet = cfg.get("查詢資料分頁", "查詢資料")
    out = run_gws(["sheets", "spreadsheets", "values", "get"],
                  params={"spreadsheetId": sid, "range": f"{sheet}!A2:G"})
    try:
        data = json.loads(out[out.index("{"):])
        vals = data.get("values", [])
    except Exception:
        print("[警告] 讀回驗證失敗（不影響已上傳資料），請自行打開試算表核對。")
        return
    remote_refund = sum(int(v[6]) for v in vals if len(v) > 6 and str(v[6]).lstrip("-").isdigit())
    local_refund = sum(r[6] for r in rows)
    ok = len(vals) == len(rows) and remote_refund == local_refund
    print(f"[驗證] 試算表列數 {len(vals)}／本地 {len(rows)}；退款總額 試算表 {remote_refund}／本地 {local_refund} → {'一致 ✅' if ok else '不一致 ❌ 請檢查！'}")
    if not ok:
        sys.exit(1)


def write_site_config(cfg, month_label):
    """產 網站\config.js（部署時一起帶上去）。logo 檔自動偵測：logo.png/jpg/jpeg/webp/svg 先找到先用。"""
    site = os.path.join(BASE, "網站")
    logo = ""
    for ext in ("png", "jpg", "jpeg", "webp", "svg"):
        if os.path.exists(os.path.join(site, f"logo.{ext}")):
            logo = f"logo.{ext}"
            break
    content = "// 這個檔由 更新斷貨查詢網站.py 自動產生，不要手改\n" + \
        "window.SITE_CONFIG = " + json.dumps({
            "apiUrl": cfg.get("apps_script_url", ""),
            "logo": logo,
            "month": month_label,
        }, ensure_ascii=False, indent=2) + ";\n"
    with open(os.path.join(site, "config.js"), "w", encoding="utf-8") as f:
        f.write(content)
    if not cfg.get("apps_script_url"):
        print("[警告] 設定.json 的 apps_script_url 還是空的——網站查詢功能不會動，請先完成 Apps Script 部署。")
    return logo


def deploy_site(cfg):
    site = os.path.join(BASE, "網站")
    proj = cfg.get("pages_project", "refund-checker")
    print(f"[info] 部署網站到 Cloudflare Pages（專案 {proj}）…")
    r = subprocess.run([_which("npx"), "--yes", "wrangler", "pages", "deploy", site,
                        "--project-name", proj, "--branch", "main", "--commit-dirty=true"],
                       capture_output=True, text=True, encoding="utf-8", errors="replace")
    out = (r.stdout or "") + (r.stderr or "")
    if r.returncode != 0:
        print(f"[錯誤] 網站部署失敗：\n{out[-800:]}", file=sys.stderr)
        print("（第一次用需要先建 Pages 專案：npx wrangler pages project create " + proj + " --production-branch main）", file=sys.stderr)
        sys.exit(1)
    m = re.search(r"https://\S+\.pages\.dev\S*", out)
    if m:
        print(f"[完成] 網站已部署：{m.group(0)}")
    else:
        print("[完成] 網站已部署（網址見上方 wrangler 輸出）")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", default=None, help="月團資料夾名稱（預設抓最新）")
    ap.add_argument("--no-deploy", action="store_true", help="只更新試算表，不部署網站")
    ap.add_argument("--no-sheet", action="store_true", help="只部署網站，不動試算表")
    ap.add_argument("--dry-run", action="store_true", help="只整理資料印統計，不上傳不部署")
    args = ap.parse_args()

    cfg = load_cfg()
    month_dir = find_month(args.month)
    month_label = re.sub(r"^\d{4}-\d{2}_", "", os.path.basename(month_dir))
    print(f"[info] 月團：{os.path.basename(month_dir)}")

    detail = find_detail_file(month_dir)
    shortage_file = find_shortage_file(month_dir)
    members = read_shipping_detail(detail)
    if not members:
        print("[錯誤] 出貨明細裡讀不到任何會員，請確認檔案內容。", file=sys.stderr)
        sys.exit(1)
    if shortage_file:
        shortages, weight_map = read_shortage(shortage_file)
    else:
        print("[警告] 找不到斷貨單，所有會員都會顯示無斷貨。")
        shortages, weight_map = {}, {}
    rows, stats = build_rows(members, shortages, weight_map, read_name_fallback(month_dir))
    print(f"[統計] 會員 {stats['members']} 位；有退款 {stats['refund_members']} 位；退款總額 {stats['refund_total']} 元")

    if args.dry_run:
        print("[dry-run] 結束，未上傳未部署。")
        return
    if not args.no_sheet:
        upload_to_sheet(cfg, rows, month_label)
        verify_sheet(cfg, rows)
    if not args.no_deploy:
        write_site_config(cfg, month_label)
        deploy_site(cfg)


if __name__ == "__main__":
    main()
