# -*- coding: utf-8 -*-
"""
匯出批次退款表.py — 把「有斷貨的會員」整理成一張批次退款總表

資料來源：Google 試算表（網站資料庫）
  查詢資料 分頁 → 誰有應退金額（斷貨單上的全部會員，含沒來填帳號的）
  退款帳號回覆 分頁 → 誰填了帳號（同會員多筆以「最新」那筆為準）

產出兩份（內容相同）：
  1. Excel：每月資料\{月團}\斷貨單\批次退款表.xlsx（就地覆寫，跑完自動同步雲端）
  2. 試算表新增/更新「批次退款總表」分頁（打開試算表就能看）

用法：
  python 匯出批次退款表.py                    # 自動抓最新月團
  python 匯出批次退款表.py --month 2026-07_7月團
  python 匯出批次退款表.py --no-upload        # 不同步雲端
"""
import argparse
import json
import os
import subprocess
import sys

sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(BASE)
sys.path.insert(0, BASE)

# 共用「更新斷貨查詢網站.py」的工具函數（設定載入、gws 呼叫、月團尋找）
import importlib.util
_spec = importlib.util.spec_from_file_location("_upd", os.path.join(BASE, "更新斷貨查詢網站.py"))
_upd = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_upd)


def fetch_values(cfg, rng):
    out = _upd.run_gws(["sheets", "spreadsheets", "values", "get"],
                       params={"spreadsheetId": cfg["spreadsheetId"], "range": rng})
    try:
        data = json.loads(out[out.index("{"):])
    except Exception:
        print(f"[錯誤] 讀不了試算表範圍 {rng}，請確認網路與 gws 登入狀態。", file=sys.stderr)
        sys.exit(1)
    return data.get("values", [])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", default=None, help="月團資料夾名稱（預設抓最新）")
    ap.add_argument("--no-upload", action="store_true", help="不自動同步雲端")
    args = ap.parse_args()

    cfg = _upd.load_cfg()
    month_dir = _upd.find_month(args.month)
    month_label = os.path.basename(month_dir)
    print(f"[info] 月團：{month_label}")

    # 1. 有應退金額的會員（來自查詢資料）
    rows = fetch_values(cfg, f"{cfg.get('查詢資料分頁','查詢資料')}!A2:G")
    refund_members = []
    for r in rows:
        r = list(r) + [""] * 7
        try:
            amt = int(str(r[6]).strip() or 0)
        except ValueError:
            amt = 0
        if amt > 0:
            refund_members.append({"id": r[0], "name": r[1], "amt": amt})
    if not refund_members:
        print("[完成] 這個月沒有任何斷貨退款會員，不需要產批次退款表。")
        return

    # 2. 已回覆（同會員取「最新」那筆；H欄類型：退款/購物金，舊資料沒有H欄視為退款）
    replies = {}
    for r in fetch_values(cfg, f"{cfg.get('回覆分頁','退款帳號回覆')}!A2:H"):
        r = list(r) + [""] * 8
        if str(r[6]).strip() == "最新" and r[1]:
            replies[str(r[1]).strip()] = {"time": r[0], "bank": str(r[4]), "acct": str(r[5]),
                                          "type": (str(r[7]).strip() or "退款")}

    # 3. 組總表
    header = ["會員編號", "姓名", "應退金額", "銀行代碼", "帳號", "填寫時間", "狀態"]
    table = []
    filled = credit = 0
    for m in refund_members:
        rep = replies.get(m["id"])
        if rep and rep["type"] == "購物金":
            credit += 1
            table.append([m["id"], m["name"], m["amt"], "", "", rep["time"], "轉購物金"])
        elif rep:
            filled += 1
            table.append([m["id"], m["name"], m["amt"], rep["bank"], rep["acct"], rep["time"], "已填"])
        else:
            table.append([m["id"], m["name"], m["amt"], "", "", "", "未填"])
    order = {"已填": 0, "轉購物金": 1, "未填": 2}
    table.sort(key=lambda x: (order.get(x[6], 9), x[0]))
    total = sum(m["amt"] for m in refund_members)
    print(f"[統計] 應退會員 {len(refund_members)} 位（已填帳號 {filled}、轉購物金 {credit}、未填 {len(refund_members)-filled-credit}）；退款總額 {total} 元")

    # 4. 寫 Excel
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "批次退款總表"
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="FFE0B2")
        c.alignment = Alignment(horizontal="center")
    for row in table:
        ws.append(row)
    for cell in ws["D"] + ws["E"]:
        cell.number_format = "@"  # 銀行代碼/帳號當文字，保住開頭的 0
    red = Font(color="D93025", bold=True)
    green = Font(color="2FA383", bold=True)
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 7).value == "未填":
            ws.cell(r, 7).font = red
        elif ws.cell(r, 7).value == "轉購物金":
            ws.cell(r, 7).font = green
    ws.append([])
    ws.append(["合計", "", total, "", "", "", ""])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.cell(ws.max_row, 3).font = Font(bold=True)
    widths = [12, 16, 12, 10, 20, 20, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    out_path = os.path.join(month_dir, "斷貨單", "批次退款表.xlsx")
    try:
        wb.save(out_path)
    except PermissionError:
        print(f"[錯誤] 無法寫入 批次退款表.xlsx：檔案正被 Excel 開著，請先關閉再重跑。", file=sys.stderr)
        sys.exit(1)
    print(f"[完成] Excel 已存：{out_path}")

    # 5. 回寫試算表「批次退款總表」分頁（沒有就建）
    sid = cfg["spreadsheetId"]
    if _upd.run_gws(["sheets", "spreadsheets", "batchUpdate"],
                    params={"spreadsheetId": sid},
                    body={"requests": [{"addSheet": {"properties": {"title": "批次退款總表"}}}]},
                    allow_fail=True) is not None:
        print("[info] 已建立 批次退款總表 分頁")
    _upd.run_gws(["sheets", "spreadsheets", "values", "clear"],
                 params={"spreadsheetId": sid, "range": "批次退款總表!A1:H"}, body={})
    values = [header] + [[str(x) for x in row] for row in table] + [[], ["合計", "", str(total)]]
    _upd.run_gws(["sheets", "spreadsheets", "values", "update"],
                 params={"spreadsheetId": sid, "range": "批次退款總表!A1", "valueInputOption": "RAW"},
                 body={"values": values})
    print("[完成] 試算表「批次退款總表」分頁已更新")

    # 6. 同步雲端
    if not args.no_upload:
        _upd_sync(out_path)


def _upd_sync(out_path):
    d = ROOT
    sync = os.path.join(d, "同步每月資料到雲端.py")
    if not os.path.exists(sync):
        return
    folder = os.path.basename(os.path.dirname(os.path.dirname(out_path)))
    print(f"[info] 同步 {folder} 到雲端…")
    r = subprocess.run([sys.executable, sync, "--only", folder], cwd=ROOT)
    if r.returncode != 0:
        print("[警告] 雲端同步失敗（本機檔案不受影響）；可稍後手動跑 同步每月資料到雲端.py")


if __name__ == "__main__":
    main()
